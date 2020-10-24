import openpyxl, random
from openpyxl import Workbook
from openpyxl.styles import Font, Color, colors
wb=Workbook()
ws=wb.active
ws.title="vznosy"

ws['A1']='Имя'
ws['B1']='Сданные профвсзносы'
ws['C1']='Месяц'
ws['D1']='Сумма'

a=int(input())
b=3*a

wb1=openpyxl.load_workbook('names.xlsx')
ws0=wb1.active
for i in range(0,a):
    ws['A'+str(i*3+2)]=ws0['A'+str(i+2)].value

#for i in range(0,a):
#    ws['A'+str(i*3+2)]='имя'+str(i+1)

for i in range(0,b):
    if i%3==0: ws['C'+str(i+2)]='сентябрь'
    if i%3==1: ws['C'+str(i+2)]='октябрь'
    if i%3==2: ws['C'+str(i+2)]='ноябрь'


for i in range(0,b):
    ws['D'+str(i+2)]=random.randint(33,35)
   

for i in range(0,b):
    if ws['D'+str(i+2)].value==35:
        ws['B'+str(i+2)]='+'
    else: ws['B'+str(i+2)]='-'



ws2=wb.create_sheet(title="SK")

ws2['A1']='Имя'
ws2['B1']='СК'
ws2['C1']='Статус'

for i in range(0,a):
    ws2['A'+str(i+2)]=ws['A'+str(i*3+2)].value

sk_list=['Кит', 'Бухта', 'Ступино']
for i in range(0,a):
    ws2['B'+str(i+2)]=random.choice(sk_list)

for i in range(0,a):
    if (ws['B'+str(i*3+2)].value=='+' and ws['B'+str(i*3+3)].value=='+' and ws['B'+str(i*3+4)].value=='+'):
        ws2['C'+str(i+2)]='Одобрено'
    else: ws2['C'+str(i+2)]='Не одобрено'

#ЗАПОЛНЕНИЕ ЗАВЕРШЕНО

for i in range(0,a):
    if (ws2['C'+str(i+2)].value=='Одобрено'):
           print(ws2['A'+str(i+2)].value, 'едет в', ws2['B'+str(i+2)].value)


wb.save('test.xlsx')
